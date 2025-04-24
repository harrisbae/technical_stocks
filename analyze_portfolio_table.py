"""
포트폴리오 분석 도구 (테이블 출력)

stocks.txt 파일에서 종목 정보를 읽어와 기술적 분석을 수행한 후 표 형식으로 출력합니다.
"""

import os
import pandas as pd
import numpy as np
import yfinance as yf
from tabulate import tabulate

def read_stocks_file(file_path='src/data/stocks.txt'):
    """
    stocks.txt 파일을 읽어 종목 정보를 리스트로 반환합니다.
    """
    stocks = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:  # 빈 줄 건너뛰기
                continue
                
            parts = line.split('/')
            if len(parts) >= 3:
                symbol = parts[0].strip()
                try:
                    price = float(parts[1].strip())
                    quantity = int(parts[2].strip())
                    stocks.append({
                        'symbol': symbol,
                        'price': price,
                        'quantity': quantity
                    })
                except ValueError:
                    print(f"경고: {line} 처리 중 오류 발생. 숫자 형식이 잘못되었습니다.")
    
    return stocks


def get_stock_data(symbol, period='1y'):
    """
    Yahoo Finance에서 종목 데이터를 가져옵니다.
    
    Parameters:
    -----------
    symbol : str
        종목 코드
    period : str
        데이터 기간 (default: '1y')
        
    Returns:
    --------
    pd.DataFrame or None
        종목 데이터 또는 실패 시 None
    """
    try:
        # Convert symbol to string if it's not already
        symbol = str(symbol).strip()
        
        # Handle empty or invalid symbols
        if not symbol:
            print("오류: 빈 종목 코드가 제공되었습니다.")
            return None
        
        # Special case for Korean stocks with names instead of ticker symbols
        if symbol == "템퍼스AI":
            symbol = "376300.KS"  # Tempus AI ticker symbol
        elif symbol == "삼성전자":
            symbol = "005930.KS"  # Samsung Electronics
        elif symbol == "현대차":
            symbol = "005380.KS"  # Hyundai Motor
        # 한국 주식인 경우 '.KS' 확장자 추가
        elif symbol.isdigit() and len(symbol) == 6:
            symbol = f"{symbol}.KS"
            
        print(f"데이터 요청 중: {symbol}")
        stock_data = yf.download(symbol, period=period, progress=False)
        
        # Validate the downloaded data
        if stock_data is None or len(stock_data) == 0:
            print(f"경고: {symbol}에 대한 데이터를 찾을 수 없습니다.")
            return None
            
        # Check if we have the necessary columns
        required_columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        if not all(col in stock_data.columns for col in required_columns):
            missing = [col for col in required_columns if col not in stock_data.columns]
            print(f"경고: {symbol} 데이터에 필요한 컬럼이 없습니다: {missing}")
            return None
            
        print(f"{symbol} 데이터 다운로드 성공: {len(stock_data)} 행")
        return stock_data
    except Exception as e:
        print(f"오류: {symbol} 데이터 다운로드 실패 - {str(e)}")
        return None


def calculate_indicators(df):
    """
    주요 기술 지표를 계산합니다.
    """
    result = {}
    
    # 종가 설정 - 명시적으로 float로 변환
    try:
        # 수정: Series의 첫 번째 요소에 명시적으로 접근
        if isinstance(df['Close'].iloc[-1], pd.Series):
            result['close'] = float(df['Close'].iloc[-1].iloc[0])
        else:
            result['close'] = float(df['Close'].iloc[-1])
    except:
        # 만약 변환에 실패하면 (Series인 경우 등) 첫 번째 요소 사용
        try:
            print(f"경고: 종가 변환에 실패했습니다. 대체 방법 시도.")
            result['close'] = float(df['Close'].iloc[-1].iloc[0])
        except:
            print(f"경고: 종가를 가져오는데 문제가 발생했습니다. 0으로 설정합니다.")
            result['close'] = 0.0
    
    # 52주 최고가/최저가 계산 (대략 1년치 데이터 기준)
    try:
        # Series에서 float 변환 이슈를 처리
        if isinstance(df['High'], pd.Series) and isinstance(df['Low'], pd.Series):
            high_values = df['High']
            low_values = df['Low']
        else:
            high_values = df['High'].iloc[:, 0]
            low_values = df['Low'].iloc[:, 0]
            
        # 최대 252 거래일(1년) 또는 가능한 최대 데이터 사용
        lookback_period = min(252, len(df))
        result['year_high'] = float(high_values[-lookback_period:].max())
        result['year_low'] = float(low_values[-lookback_period:].min())
        
        # 박스권 계산
        current_price = result['close']
        box_range = result['year_high'] - result['year_low']
        
        if box_range > 0:
            result['box_position'] = round(((current_price - result['year_low']) / box_range) * 100, 2)
        else:
            result['box_position'] = 50.0
            
        # 매수/매도 가격 추천
        # 단순 박스권 기반 계산 (예: 하위 20%는 매수, 상위 20%는 매도)
        result['buy_price'] = round(result['year_low'] + (box_range * 0.2), 2)
        result['sell_price'] = round(result['year_high'] - (box_range * 0.2), 2)
        
    except Exception as e:
        print(f"52주 최고/최저가 계산 중 오류: {e}")
        result['year_high'] = result['close'] * 1.2  # 임의 설정
        result['year_low'] = result['close'] * 0.8   # 임의 설정
        result['box_position'] = 50.0
        result['buy_price'] = result['close'] * 0.9
        result['sell_price'] = result['close'] * 1.1
    
    # 볼린저 밴드 계산
    try:
        period = 20
        if len(df) < period:
            result['boll_position'] = 50.0
            result['boll_eval'] = '데이터 부족'
        else:
            # Series에서 float 변환 이슈를 처리
            if isinstance(df['Close'], pd.Series):
                close_values = df['Close']
            else:
                close_values = df['Close'].iloc[:, 0]
                
            # 중심선 (20일 이동평균)
            middle_band = close_values.rolling(window=period).mean().iloc[-1]
            
            # 표준편차 계산
            std = close_values.rolling(window=period).std().iloc[-1]
            
            # Series 변환 처리
            if isinstance(middle_band, pd.Series):
                middle_band = float(middle_band.iloc[0])
                std = float(std.iloc[0])
            else:
                middle_band = float(middle_band)
                std = float(std)
            
            # 상하단 밴드
            upper_band = middle_band + 2 * std
            lower_band = middle_band - 2 * std
            
            # 밴드 범위와 위치 계산
            band_range = upper_band - lower_band
            
            if np.isclose(band_range, 0):
                band_position = 50.0
            else:
                band_position = ((result['close'] - lower_band) / band_range) * 100
                
            result['boll_position'] = round(band_position, 2)
            
            # 밴드 위치에 따른 평가
            if band_position > 80:
                result['boll_eval'] = '매도'
            elif band_position < 20:
                result['boll_eval'] = '매수'
            else:
                result['boll_eval'] = '중립'
    except Exception as e:
        print(f"볼린저 밴드 계산 중 오류: {e}")
        result['boll_position'] = 50.0
        result['boll_eval'] = '오류'
        
    # MACD 계산
    try:
        short_window = 12
        long_window = 26
        signal_window = 9
        
        if len(df) < max(long_window, signal_window):
            result['macd_value'] = 0.0
            result['macd_signal'] = 0.0
            result['macd_diff'] = 0.0
            result['macd_eval'] = '데이터 부족'
        else:
            # Series에서 float 변환 이슈를 처리
            if isinstance(df['Close'], pd.Series):
                close_values = df['Close']
            else:
                close_values = df['Close'].iloc[:, 0]
                
            short_ema = close_values.ewm(span=short_window, adjust=False).mean()
            long_ema = close_values.ewm(span=long_window, adjust=False).mean()
            
            macd = short_ema - long_ema
            signal = macd.ewm(span=signal_window, adjust=False).mean()
            
            if isinstance(macd.iloc[-1], pd.Series):
                macd_value = float(macd.iloc[-1].iloc[0])
                signal_value = float(signal.iloc[-1].iloc[0])
            else:
                macd_value = float(macd.iloc[-1])
                signal_value = float(signal.iloc[-1])
            
            result['macd_value'] = macd_value
            result['macd_signal'] = signal_value
            
            macd_diff = result['macd_value'] - result['macd_signal']
            result['macd_diff'] = macd_diff
            
            # MACD 평가
            if macd_diff > 0 and result['macd_value'] > 0:
                result['macd_eval'] = '강매수'
            elif macd_diff > 0:
                result['macd_eval'] = '약매수'
            elif macd_diff < 0 and result['macd_value'] < 0:
                result['macd_eval'] = '강매도'
            else:
                result['macd_eval'] = '약매도'
    except Exception as e:
        print(f"MACD 계산 중 오류: {e}")
        result['macd_value'] = 0.0
        result['macd_signal'] = 0.0
        result['macd_diff'] = 0.0
        result['macd_eval'] = '오류'
        
    # RSI 계산
    try:
        period = 14
        if len(df) < period + 1:
            result['rsi'] = 50.0
            result['rsi_eval'] = '데이터 부족'
        else:
            # Series에서 float 변환 이슈를 처리
            if isinstance(df['Close'], pd.Series):
                close_values = df['Close']
            else:
                close_values = df['Close'].iloc[:, 0]
                
            delta = close_values.diff(1)
            gain = delta.copy()
            loss = delta.copy()
            
            gain[gain < 0] = 0
            loss[loss > 0] = 0
            loss = abs(loss)
            
            avg_gain_value = gain.rolling(window=period).mean().iloc[-1]
            avg_loss_value = loss.rolling(window=period).mean().iloc[-1]
            
            # Series 변환 처리
            if isinstance(avg_gain_value, pd.Series):
                avg_gain_value = float(avg_gain_value.iloc[0])
                avg_loss_value = float(avg_loss_value.iloc[0])
            else:
                avg_gain_value = float(avg_gain_value)
                avg_loss_value = float(avg_loss_value)
            
            if avg_loss_value == 0:
                rsi_value = 100.0
            else:
                rs = avg_gain_value / avg_loss_value
                rsi_value = 100 - (100 / (1 + rs))
                
            result['rsi'] = rsi_value
            
            # RSI 평가
            if result['rsi'] > 70:
                result['rsi_eval'] = '과매수'
            elif result['rsi'] < 30:
                result['rsi_eval'] = '과매도'
            else:
                result['rsi_eval'] = '중립'
    except Exception as e:
        print(f"RSI 계산 중 오류: {e}")
        result['rsi'] = 50.0
        result['rsi_eval'] = '오류'
        
    # CCI 계산 (Commodity Channel Index)
    try:
        period = 20
        if len(df) < period:
            result['cci'] = 0.0
            result['cci_eval'] = '데이터 부족'
        else:
            # Series에서 DataFrame 변환 처리
            if isinstance(df['High'], pd.Series) and isinstance(df['Low'], pd.Series) and isinstance(df['Close'], pd.Series):
                high_values = df['High']
                low_values = df['Low']
                close_values = df['Close']
            else:
                high_values = df['High'].iloc[:, 0]
                low_values = df['Low'].iloc[:, 0]
                close_values = df['Close'].iloc[:, 0]
            
            tp = (high_values + low_values + close_values) / 3
            tp_ma = tp.rolling(window=period).mean()
            md = tp.rolling(window=period).apply(lambda x: np.abs(x - x.mean()).mean())
            
            tp_last = tp.iloc[-1]
            tp_ma_last = tp_ma.iloc[-1]
            md_last = md.iloc[-1]
            
            # Series 변환 처리
            if isinstance(tp_last, pd.Series):
                tp_last = float(tp_last.iloc[0])
                tp_ma_last = float(tp_ma_last.iloc[0])
                md_last = float(md_last.iloc[0])
            
            if md_last == 0:
                cci = 0
            else:
                cci = (tp_last - tp_ma_last) / (0.015 * md_last)
                
            result['cci'] = cci
            
            # CCI 평가
            if result['cci'] > 100:
                result['cci_eval'] = '과매수'
            elif result['cci'] < -100:
                result['cci_eval'] = '과매도'
            else:
                result['cci_eval'] = '중립'
    except Exception as e:
        print(f"CCI 계산 중 오류: {e}")
        result['cci'] = 0.0
        result['cci_eval'] = '오류'
        
    # ADX 계산 (Average Directional Index) - 추가
    try:
        period = 14
        if len(df) < period + 1:
            result['adx'] = 0.0
            result['pdi'] = 0.0
            result['mdi'] = 0.0
            result['adx_eval'] = '데이터 부족'
        else:
            # Series에서 DataFrame 변환 처리
            if isinstance(df['High'], pd.Series) and isinstance(df['Low'], pd.Series) and isinstance(df['Close'], pd.Series):
                high_values = df['High']
                low_values = df['Low']
                close_values = df['Close']
            else:
                high_values = df['High'].iloc[:, 0]
                low_values = df['Low'].iloc[:, 0]
                close_values = df['Close'].iloc[:, 0]
            
            # 가격 이동 계산
            high_diff = high_values.diff(1)
            low_diff = low_values.diff(1).multiply(-1)
            
            # True Range 계산
            tr1 = high_values - low_values
            tr2 = (high_values - close_values.shift(1)).abs()
            tr3 = (low_values - close_values.shift(1)).abs()
            tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
            
            # +DM, -DM 계산
            plus_dm = pd.Series(0.0, index=df.index)
            minus_dm = pd.Series(0.0, index=df.index)
            
            for i in range(1, len(df)):
                if high_diff.iloc[i] > low_diff.iloc[i] and high_diff.iloc[i] > 0:
                    plus_dm.iloc[i] = high_diff.iloc[i]
                if low_diff.iloc[i] > high_diff.iloc[i] and low_diff.iloc[i] > 0:
                    minus_dm.iloc[i] = low_diff.iloc[i]
            
            # TR, +DM, -DM의 평활화 (Smoothed average)
            atr = tr.ewm(alpha=1/period, adjust=False).mean()
            plus_di = 100 * (plus_dm.ewm(alpha=1/period, adjust=False).mean() / atr)
            minus_di = 100 * (minus_dm.ewm(alpha=1/period, adjust=False).mean() / atr)
            
            # DX 계산
            dx = 100 * ((plus_di - minus_di).abs() / (plus_di + minus_di).abs())
            
            # ADX 계산
            adx = dx.ewm(alpha=1/period, adjust=False).mean()
            
            # 마지막 값 가져오기
            adx_value = float(adx.iloc[-1])
            pdi_value = float(plus_di.iloc[-1])
            mdi_value = float(minus_di.iloc[-1])
            
            # Series 변환 처리
            if isinstance(adx_value, pd.Series):
                adx_value = float(adx_value.iloc[0])
                pdi_value = float(pdi_value.iloc[0])
                mdi_value = float(mdi_value.iloc[0])
            
            result['adx'] = adx_value
            result['pdi'] = pdi_value
            result['mdi'] = mdi_value
            
            # ADX 평가
            if adx_value < 20:
                result['adx_eval'] = '약한추세'
            elif adx_value < 40:
                if pdi_value > mdi_value:
                    result['adx_eval'] = '상승추세'
                else:
                    result['adx_eval'] = '하락추세'
            else:
                if pdi_value > mdi_value:
                    result['adx_eval'] = '강한상승'
                else:
                    result['adx_eval'] = '강한하락'
    except Exception as e:
        print(f"ADX 계산 중 오류: {e}")
        result['adx'] = 0.0
        result['pdi'] = 0.0
        result['mdi'] = 0.0
        result['adx_eval'] = '오류'
        
    # Williams %R 계산 - 추가
    try:
        period = 14
        if len(df) < period:
            result['williams'] = -50.0
            result['williams_eval'] = '데이터 부족'
        else:
            # Series에서 DataFrame 변환 처리
            if isinstance(df['High'], pd.Series) and isinstance(df['Low'], pd.Series) and isinstance(df['Close'], pd.Series):
                high_values = df['High']
                low_values = df['Low']
                close_values = df['Close']
            else:
                high_values = df['High'].iloc[:, 0]
                low_values = df['Low'].iloc[:, 0]
                close_values = df['Close'].iloc[:, 0]
            
            # 기간 내 최고가 및 최저가
            highest_high = high_values.rolling(window=period).max()
            lowest_low = low_values.rolling(window=period).min()
            
            # Williams %R 계산
            williams_r = ((highest_high - close_values) / (highest_high - lowest_low)) * -100
            
            # 마지막 값 가져오기
            williams_value = float(williams_r.iloc[-1])
            
            # Series 변환 처리
            if isinstance(williams_value, pd.Series):
                williams_value = float(williams_value.iloc[0])
            
            result['williams'] = williams_value
            
            # Williams %R 평가
            if williams_value > -20:
                result['williams_eval'] = '과매수'
            elif williams_value < -80:
                result['williams_eval'] = '과매도'
            else:
                result['williams_eval'] = '중립'
    except Exception as e:
        print(f"Williams %R 계산 중 오류: {e}")
        result['williams'] = -50.0
        result['williams_eval'] = '오류'
        
    # Stochastic 계산 - 추가
    try:
        k_period = 14
        d_period = 3
        if len(df) < k_period:
            result['stoch_k'] = 50.0
            result['stoch_d'] = 50.0
            result['stoch_eval'] = '데이터 부족'
        else:
            # Series에서 DataFrame 변환 처리
            if isinstance(df['High'], pd.Series) and isinstance(df['Low'], pd.Series) and isinstance(df['Close'], pd.Series):
                high_values = df['High']
                low_values = df['Low']
                close_values = df['Close']
            else:
                high_values = df['High'].iloc[:, 0]
                low_values = df['Low'].iloc[:, 0]
                close_values = df['Close'].iloc[:, 0]
            
            # 기간 내 최고가 및 최저가
            highest_high = high_values.rolling(window=k_period).max()
            lowest_low = low_values.rolling(window=k_period).min()
            
            # %K 계산
            stoch_k = ((close_values - lowest_low) / (highest_high - lowest_low)) * 100
            
            # %D 계산 (k의 d_period 이동평균)
            stoch_d = stoch_k.rolling(window=d_period).mean()
            
            # 마지막 값 가져오기
            k_value = float(stoch_k.iloc[-1])
            d_value = float(stoch_d.iloc[-1])
            
            # Series 변환 처리
            if isinstance(k_value, pd.Series):
                k_value = float(k_value.iloc[0])
                d_value = float(d_value.iloc[0])
            
            result['stoch_k'] = k_value
            result['stoch_d'] = d_value
            
            # Stochastic 평가
            if k_value > 80:
                if k_value > d_value:
                    result['stoch_eval'] = '과매수반전'
                else:
                    result['stoch_eval'] = '과매수하락'
            elif k_value < 20:
                if k_value > d_value:
                    result['stoch_eval'] = '과매도반전'
                else:
                    result['stoch_eval'] = '과매도하락'
            else:
                if k_value > d_value:
                    result['stoch_eval'] = '상승추세'
                else:
                    result['stoch_eval'] = '하락추세'
    except Exception as e:
        print(f"Stochastic 계산 중 오류: {e}")
        result['stoch_k'] = 50.0
        result['stoch_d'] = 50.0
        result['stoch_eval'] = '오류'
    
    return result


def analyze_stock(symbol, stock_data, purchase_price):
    """
    종목 데이터를 분석합니다.
    """
    # 기술적 지표 계산
    indicators = calculate_indicators(stock_data)
    
    # 현재 가격
    current_price = indicators['close']
    
    # Series 객체인 경우 스칼라 값으로 변환
    if isinstance(current_price, pd.Series):
        current_price = float(current_price.iloc[0])
    
    # 수익률 계산
    if purchase_price:
        profit_percentage = ((current_price - purchase_price) / purchase_price) * 100
        profit_amount = current_price - purchase_price
    else:
        profit_percentage = None
        profit_amount = None
    
    # Series 객체인 경우 스칼라 값으로 변환
    for key in ['boll_position', 'macd_diff', 'rsi', 'cci', 'adx', 'williams', 'stoch_k', 'stoch_d',
                'year_high', 'year_low', 'box_position', 'buy_price', 'sell_price']:
        if key in indicators and isinstance(indicators[key], pd.Series):
            indicators[key] = float(indicators[key].iloc[0])
    
    # 종합평가 계산
    overall_eval = calculate_overall_evaluation(indicators)
    
    # 결과 반환
    return {
        'symbol': symbol,
        'purchase_price': purchase_price,
        'current_price': current_price,
        'profit_percentage': profit_percentage,
        'profit_amount': profit_amount,
        'year_high': indicators.get('year_high'),
        'year_low': indicators.get('year_low'),
        'box_position': indicators.get('box_position'),
        'buy_price': indicators.get('buy_price'),
        'sell_price': indicators.get('sell_price'),
        'boll_position': indicators.get('boll_position'),
        'boll_eval': indicators.get('boll_eval'),
        'macd_diff': indicators.get('macd_diff'),
        'macd_eval': indicators.get('macd_eval'),
        'rsi': indicators.get('rsi'),
        'rsi_eval': indicators.get('rsi_eval'),
        'cci': indicators.get('cci'),
        'cci_eval': indicators.get('cci_eval'),
        'adx': indicators.get('adx'),
        'pdi': indicators.get('pdi'),
        'mdi': indicators.get('mdi'),
        'adx_eval': indicators.get('adx_eval'),
        'williams': indicators.get('williams'),
        'williams_eval': indicators.get('williams_eval'),
        'stoch_k': indicators.get('stoch_k'),
        'stoch_d': indicators.get('stoch_d'),
        'stoch_eval': indicators.get('stoch_eval'),
        'overall_eval': overall_eval  # 종합평가 추가
    }

def calculate_overall_evaluation(indicators):
    """
    지표간 상관관계를 기반으로 종합평가를 수행합니다.
    7대 주요 지표(볼린저, MACD, RSI, CCI, ADX, Williams %R, Stochastic)를 분석하여
    강매수/매수/중립/매도/강매도 신호를 결정합니다.
    """
    # 각 지표의 평가 점수화
    scores = {
        'buy': 0,
        'neutral': 0,
        'sell': 0
    }
    
    # 지표값 추출
    boll_position = indicators.get('boll_position', 50)
    boll_eval = indicators.get('boll_eval', '중립')
    macd_diff = indicators.get('macd_diff', 0)
    macd_eval = indicators.get('macd_eval', '중립')
    rsi = indicators.get('rsi', 50)
    rsi_eval = indicators.get('rsi_eval', '중립')
    cci = indicators.get('cci', 0)
    cci_eval = indicators.get('cci_eval', '중립')
    adx = indicators.get('adx', 0)
    pdi = indicators.get('pdi', 0)
    mdi = indicators.get('mdi', 0)
    adx_eval = indicators.get('adx_eval', '약한추세')
    williams = indicators.get('williams', -50)
    williams_eval = indicators.get('williams_eval', '중립')
    stoch_k = indicators.get('stoch_k', 50)
    stoch_d = indicators.get('stoch_d', 50)
    stoch_eval = indicators.get('stoch_eval', '중립')
    
    # 1. 볼린저 밴드 (변동성 기반)
    if boll_position < 20:
        scores['buy'] += 2  # 하단 근접/이탈 시 더 강한 매수 신호
    elif boll_position < 40:
        scores['buy'] += 1
    elif boll_position > 80:
        scores['sell'] += 2  # 상단 근접/돌파 시 더 강한 매도 신호
    elif boll_position > 60:
        scores['sell'] += 1
    else:
        scores['neutral'] += 1
    
    # 2. MACD (추세 기반)
    if macd_eval == '강매수':
        scores['buy'] += 2
    elif macd_eval == '약매수':
        scores['buy'] += 1
    elif macd_eval == '강매도':
        scores['sell'] += 2
    elif macd_eval == '약매도':
        scores['sell'] += 1
    else:
        scores['neutral'] += 1
    
    # 3. RSI (모멘텀 기반)
    if rsi < 30:
        scores['buy'] += 2  # 30 이하는 강한 과매도 신호
    elif rsi < 45:
        scores['buy'] += 1  # 30-45 구간은 약한 매수 신호
    elif rsi > 70:
        scores['sell'] += 2  # 70 이상은 강한 과매수 신호
    elif rsi > 60:
        scores['sell'] += 1  # 60-70 구간은 약한 매도 신호
    else:
        scores['neutral'] += 1  # 45-60 구간은 중립
    
    # 4. CCI (사이클 기반)
    if cci < -100:
        scores['buy'] += 2  # -100 이하는 강한 매수 신호
    elif cci < 0:
        scores['buy'] += 1
    elif cci > 100:
        scores['sell'] += 2  # 100 이상은 강한 매도 신호
    elif cci > 0:
        scores['sell'] += 1
    else:
        scores['neutral'] += 1
    
    # 5. ADX (추세 강도)
    # ADX > 25는 강한 추세, 방향은 +DI와 -DI로 판단
    if adx > 25:
        if pdi > mdi:  # 상승 추세
            scores['buy'] += 2
        else:  # 하락 추세
            scores['sell'] += 2
    elif adx > 20:  # 약한 추세 형성 중
        if pdi > mdi:
            scores['buy'] += 1
        else:
            scores['sell'] += 1
    else:
        scores['neutral'] += 1  # 추세 없음 (ADX < 20)
    
    # 6. Williams %R (모멘텀)
    if williams < -80:
        scores['buy'] += 2  # -80 이하는 강한 과매도 신호
    elif williams < -50:
        scores['buy'] += 1
    elif williams > -20:
        scores['sell'] += 2  # -20 이상은 강한 과매수 신호
    elif williams > -50:
        scores['sell'] += 1
    else:
        scores['neutral'] += 1
    
    # 7. Stochastic (모멘텀)
    # %K와 %D의 관계 및 위치를 모두 고려
    if stoch_k < 20:  # 과매도 구간
        if stoch_k > stoch_d:  # %K가 %D를 상향돌파
            scores['buy'] += 2  # 강한 매수 신호
        else:
            scores['buy'] += 1
    elif stoch_k > 80:  # 과매수 구간
        if stoch_k < stoch_d:  # %K가 %D를 하향돌파
            scores['sell'] += 2  # 강한 매도 신호
        else:
            scores['sell'] += 1
    else:
        if stoch_k > stoch_d:  # 중립 구간에서 %K가 %D보다 위에 있으면 상승 추세
            scores['buy'] += 1
        elif stoch_k < stoch_d:  # 중립 구간에서 %K가 %D보다 아래에 있으면 하락 추세
            scores['sell'] += 1
        else:
            scores['neutral'] += 1
    
    # ----- 지표 간 상관관계 기반 추가 점수 부여 -----
    
    # 1. MACD + ADX 조합 (추세 방향 + 강도)
    if macd_diff > 0 and adx > 25 and pdi > mdi:
        scores['buy'] += 3  # 강한 상승 추세 확인 시 추가 점수
    elif macd_diff < 0 and adx > 25 and pdi < mdi:
        scores['sell'] += 3  # 강한 하락 추세 확인 시 추가 점수
    
    # 2. 모멘텀 지표 조합 (RSI + Williams %R + CCI)
    momentum_buy = 0
    momentum_sell = 0
    
    # RSI 체크
    if rsi < 30:
        momentum_buy += 1
    elif rsi > 70:
        momentum_sell += 1
    
    # Williams %R 체크
    if williams < -80:
        momentum_buy += 1
    elif williams > -20:
        momentum_sell += 1
    
    # CCI 체크
    if cci < -100:
        momentum_buy += 1
    elif cci > 100:
        momentum_sell += 1
    
    # 3개 중 2개 이상이 같은 신호를 보이면 강한 시그널로 판단
    if momentum_buy >= 2:
        scores['buy'] += 3
    if momentum_sell >= 2:
        scores['sell'] += 3
    
    # 3. 볼린저 밴드 + Stochastic 조합 (변동성 + 모멘텀)
    if boll_position < 20 and stoch_k < 20 and stoch_k > stoch_d:
        scores['buy'] += 2  # 밴드 하단 + Stochastic 상향돌파 = 강한 반등 신호
    if boll_position > 80 and stoch_k > 80 and stoch_k < stoch_d:
        scores['sell'] += 2  # 밴드 상단 + Stochastic 하향돌파 = 강한 하락 신호
    
    # 4. CCI + Stochastic 조합 (사이클 + 모멘텀)
    if cci < -100 and stoch_k < 20 and stoch_k > stoch_d:
        scores['buy'] += 2  # 두 지표 모두 과매도에서 반등 신호
    if cci > 100 and stoch_k > 80 and stoch_k < stoch_d:
        scores['sell'] += 2  # 두 지표 모두 과매수에서 하락 신호
    
    # ------ 종합평가 결정 ------
    
    # 예시 케이스에 맞는 종합 점수 평가
    total_score = scores['buy'] + scores['neutral'] + scores['sell']
    buy_percentage = (scores['buy'] / total_score) * 100
    sell_percentage = (scores['sell'] / total_score) * 100
    
    # 평가 결정
    # 실전 전략 적용 - 제공된 케이스 기준으로 평가
    evaluation = ""
    score_detail = f"({scores['buy']}:{scores['neutral']}:{scores['sell']})"
    
    # 종합평가에 추가할 근거 문자열 초기화
    reasoning = []
    
    # 매우 강한 신호들 확인 (근거에 추가)
    if momentum_buy >= 2:
        reasoning.append("모멘텀 지표 매수신호")
    if momentum_sell >= 2:
        reasoning.append("모멘텀 지표 매도신호")
    
    # 추가 조합 근거
    if macd_diff > 0 and adx > 25 and pdi > mdi:
        reasoning.append("강한상승추세")
    elif macd_diff < 0 and adx > 25 and pdi < mdi:
        reasoning.append("강한하락추세")
    
    # 강매수 케이스 - buy 점수가 높고, 핵심 조건 충족
    if scores['buy'] > scores['sell'] * 2 and buy_percentage >= 65:
        # 추가 확인: 주요 지표 다수가 강매수 신호
        if ((rsi < 30 or williams < -80 or cci < -100) and  # 과매도 상태
            (stoch_k < 30 and stoch_k > stoch_d) and  # Stochastic 상향돌파
            (adx > 25 and pdi > mdi)):  # 강한 상승 추세
            reasoning.append("과매도 반등신호")
            evaluation = f"강매수 {score_detail}"
        else:
            evaluation = f"매수 {score_detail}"
    
    # 매수 케이스 - buy 점수가 우세하지만 강매수보다는 약함
    elif scores['buy'] > scores['sell'] and buy_percentage >= 45:
        evaluation = f"매수 {score_detail}"
    
    # 강매도 케이스 - sell 점수가 높고, 핵심 조건 충족
    elif scores['sell'] > scores['buy'] * 2 and sell_percentage >= 65:
        # 추가 확인: 주요 지표 다수가 강매도 신호
        if ((rsi > 70 or williams > -20 or cci > 100) and  # 과매수 상태
            (stoch_k > 70 and stoch_k < stoch_d) and  # Stochastic 하향돌파
            (adx > 25 and pdi < mdi)):  # 강한 하락 추세
            reasoning.append("과매수 조정신호")
            evaluation = f"강매도 {score_detail}"
        else:
            evaluation = f"매도 {score_detail}"
    
    # 매도 케이스 - sell 점수가 우세하지만 강매도보다는 약함
    elif scores['sell'] > scores['buy'] and sell_percentage >= 45:
        evaluation = f"매도 {score_detail}"
    
    # 중립/보유 케이스 - 명확한 방향성이 없음
    else:
        evaluation = f"보유 {score_detail}"
    
    # 근거 추가 (최대 3개까지)
    if reasoning:
        evaluation += f" - {', '.join(reasoning[:3])}"
    
    return evaluation


def analyze_portfolio(stocks):
    """
    포트폴리오 내 모든 종목을 분석합니다.
    """
    results = []
    total_investment = 0.0
    total_current_value = 0.0
    
    print(f"총 {len(stocks)}개 종목 분석 시작...\n")
    
    for stock in stocks:
        symbol = stock['symbol']
        purchase_price = stock['price']
        quantity = stock['quantity']
        
        print(f"{symbol} 분석 중...")
        
        # 종목 데이터 가져오기
        stock_data = get_stock_data(symbol)
        if stock_data is None:
            continue
            
        # 종목 분석
        analysis = analyze_stock(symbol, stock_data, purchase_price)
        
        # 수량 정보 추가
        analysis['quantity'] = quantity
        
        # 투자금액과 현재가치 계산
        investment = purchase_price * quantity
        current_value = analysis['current_price'] * quantity
        
        # Series 객체인 경우 스칼라 값으로 변환
        if isinstance(investment, pd.Series):
            investment = float(investment.iloc[0])
        if isinstance(current_value, pd.Series):
            current_value = float(current_value.iloc[0])
        
        analysis['investment'] = investment
        analysis['current_value'] = current_value
        
        # 총계에 추가
        total_investment += float(investment)
        total_current_value += float(current_value)
        
        results.append(analysis)
    
    # 포트폴리오 요약 정보
    summary = {
        'total_investment': total_investment,
        'total_current_value': total_current_value,
        'profit_percentage': ((total_current_value - total_investment) / total_investment * 100) if total_investment > 0 else 0,
        'profit_amount': total_current_value - total_investment
    }
    
    return results, summary


def display_results_table(results, summary):
    """
    분석 결과를 표 형태로 출력하고 엑셀 파일로 저장합니다.
    """
    # 데이터 형식 확인 및 변환
    formatted_results = []
    excel_results = []  # 엑셀 출력용 데이터
    
    for r in results:
        try:
            # Series 객체 처리
            for key in r:
                if isinstance(r[key], pd.Series):
                    r[key] = float(r[key].iloc[0])
            
            # USD 가격 필드는 소수점 2자리까지 반올림 처리
            r['purchase_price'] = round(r['purchase_price'], 2) if r['purchase_price'] is not None else None
            r['current_price'] = round(r['current_price'], 2) if r['current_price'] is not None else None
            r['profit_amount'] = round(r['profit_amount'], 2) if r['profit_amount'] is not None else None
            r['year_high'] = round(r['year_high'], 2) if r['year_high'] is not None else None
            r['year_low'] = round(r['year_low'], 2) if r['year_low'] is not None else None
            r['buy_price'] = round(r['buy_price'], 2) if r['buy_price'] is not None else None
            r['sell_price'] = round(r['sell_price'], 2) if r['sell_price'] is not None else None
            
            # 수익률은 소수점 2자리까지 반올림
            if r['profit_percentage'] is not None:
                r['profit_percentage'] = round(r['profit_percentage'], 2)
            
            # 수익률 계산
            profit_str = f"{r['profit_percentage']:.2f}% (${r['profit_amount']:,.2f})" if r['profit_percentage'] is not None else "N/A"
            
            # 박스권 정보
            box_str = f"{r['box_position']:.2f}% (${r['year_low']:,.2f}-${r['year_high']:,.2f})" if r['box_position'] is not None else "N/A"
            buy_sell_str = f"매수: ${r['buy_price']:,.2f} / 매도: ${r['sell_price']:,.2f}" if r['buy_price'] is not None else "N/A"
            
            # 기술적 지표 평가
            boll_str = f"{r['boll_position']:.2f} ({r['boll_eval']})" if r['boll_position'] is not None else "N/A"
            macd_str = f"{r['macd_diff']:.2f} ({r['macd_eval']})" if r['macd_diff'] is not None else "N/A"
            rsi_str = f"{r['rsi']:.2f} ({r['rsi_eval']})" if r['rsi'] is not None else "N/A"
            cci_str = f"{r['cci']:.2f} ({r['cci_eval']})" if r['cci'] is not None else "N/A"
            
            # 새로 추가된 지표
            adx_str = f"{r['adx']:.2f} ({r['adx_eval']})" if 'adx' in r and r['adx'] is not None else "N/A"
            williams_str = f"{r['williams']:.2f} ({r['williams_eval']})" if 'williams' in r and r['williams'] is not None else "N/A"
            stoch_str = f"K:{r['stoch_k']:.2f} D:{r['stoch_d']:.2f} ({r['stoch_eval']})" if 'stoch_k' in r and r['stoch_k'] is not None else "N/A"
            
            # 종합평가
            overall_str = r['overall_eval'] if 'overall_eval' in r else "N/A"
            
            # 결과 형식화
            formatted_row = [
                r['symbol'],
                f"{r['quantity']}주",
                f"${r['purchase_price']:,.2f}",
                f"${r['current_price']:,.2f}",
                profit_str,
                box_str,
                buy_sell_str,
                overall_str,  # 종합평가 위치 이동
                boll_str,
                macd_str,
                rsi_str,
                cci_str,
                adx_str,
                williams_str,
                stoch_str
            ]
            formatted_results.append(formatted_row)
            
            # 엑셀 출력용 데이터 구성
            excel_row = {
                "종목": r['symbol'],
                "수량": r['quantity'],
                "매수가": r['purchase_price'],
                "현재가": r['current_price'],
                "수익률(%)": r['profit_percentage'] if r['profit_percentage'] is not None else 0,
                "수익금액($)": r['profit_amount'] if r['profit_amount'] is not None else 0,
                "52주 최저가($)": r['year_low'],
                "52주 최고가($)": r['year_high'],
                "박스권 위치(%)": r['box_position'],
                "추천 매수가($)": r['buy_price'],
                "추천 매도가($)": r['sell_price'],
                "종합평가": r['overall_eval'] if r['overall_eval'] else "N/A",
                "볼린저 위치": r['boll_position'],
                "볼린저 평가": r['boll_eval'],
                "MACD 값": r['macd_diff'],
                "MACD 평가": r['macd_eval'],
                "RSI": r['rsi'],
                "RSI 평가": r['rsi_eval'],
                "CCI": r['cci'],
                "CCI 평가": r['cci_eval'],
                "ADX": r['adx'],
                "ADX 평가": r['adx_eval'],
                "Williams %R": r['williams'],
                "Williams 평가": r['williams_eval'],
                "Stochastic %K": r['stoch_k'],
                "Stochastic %D": r['stoch_d'],
                "Stochastic 평가": r['stoch_eval']
            }
            excel_results.append(excel_row)
            
        except Exception as e:
            print(f"Error formatting result for {r['symbol']}: {e}")
    
    # 테이블 헤더
    headers = ["종목", "수량", "매수가", "현재가", "수익률", "박스권", "추천가격", "종합평가", "볼린저", "MACD", "RSI", "CCI", "ADX", "Williams %R", "Stochastic"]
    
    # 결과 출력
    print(tabulate(formatted_results, headers=headers, tablefmt="grid"))
    
    # 요약 정보 출력 (소수점 2자리로 표시)
    summary['profit_percentage'] = round(summary['profit_percentage'], 2) if summary['profit_percentage'] is not None else 0
    # 금액 관련 필드 반올림 처리
    summary['total_investment'] = round(summary['total_investment'], 2)
    summary['total_current_value'] = round(summary['total_current_value'], 2)
    summary['profit_amount'] = round(summary['profit_amount'], 2)
    
    print("\n포트폴리오 요약:")
    print(f"총 투자금액: ${summary['total_investment']:,.2f}")
    print(f"현재 가치: ${summary['total_current_value']:,.2f}")
    print(f"전체 수익률: {summary['profit_percentage']:.2f}% (${summary['profit_amount']:,.2f})")
    
    # 엑셀 파일로 저장
    try:
        # 데이터프레임 생성
        df = pd.DataFrame(excel_results)
        
        # 폴더 생성 (존재하지 않는 경우)
        os.makedirs('temp', exist_ok=True)
        
        # 현재 시간을 파일명에 포함
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"temp/portfolio_analysis_{timestamp}.xlsx"
        csv_filename = f"temp/portfolio_analysis_{timestamp}.csv"
        
        # 요약 정보 추가 (소수점 2자리로 표시)
        summary_df = pd.DataFrame({
            '항목': ['총 투자금액', '현재 가치', '전체 수익률', '수익금액'],
            '값': [
                f"${summary['total_investment']:,.2f}", 
                f"${summary['total_current_value']:,.2f}", 
                f"{summary['profit_percentage']:.2f}%", 
                f"${summary['profit_amount']:,.2f}"
            ]
        })
        
        # Excel 파일로 저장 (셀 색상 지정)
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # 데이터프레임을 엑셀에 저장
            df.to_excel(writer, sheet_name='포트폴리오 분석', index=False)
            summary_df.to_excel(writer, sheet_name='요약', index=False)
            
            # 워크북과 워크시트 가져오기
            workbook = writer.book
            worksheet = writer.sheets['포트폴리오 분석']
            
            # 색상 정의
            red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')  # 연한 빨간색
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # 연한 녹색
            
            # 종합평가 열의 인덱스 찾기
            column_idx = None
            for idx, cell in enumerate(worksheet[1]):
                if cell.value == '종합평가':
                    column_idx = idx + 1  # 1부터 시작하는 엑셀 열 인덱스
                    break
            
            if column_idx:
                # 모든 행 순회하면서 셀 색상 지정
                for row_idx in range(2, worksheet.max_row + 1):  # 헤더를 제외한 데이터 행
                    cell = worksheet.cell(row=row_idx, column=column_idx)
                    
                    if cell.value and isinstance(cell.value, str):
                        if cell.value.startswith('매수') or cell.value.startswith('강매수'):
                            cell.fill = green_fill
                        elif cell.value.startswith('매도') or cell.value.startswith('강매도'):
                            cell.fill = red_fill
            
            # 열 너비 자동 조정
            for idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = None
                
                # 각 셀의 내용 길이 확인하여 최대 길이 찾기
                for cell in column:
                    try:
                        # merged cell 건너뛰기
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                                
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # 열 너비 설정 (약간의 여유 공간 추가)
                if column_letter:
                    adjusted_width = (max_length + 2) * 1.0
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 요약 시트도 열 너비 조정
            summary_sheet = writer.sheets['요약']
            for idx, column in enumerate(summary_sheet.columns, 1):
                max_length = 0
                column_letter = None
                
                for cell in column:
                    try:
                        # merged cell 건너뛰기
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                                
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                title_length = len(str(column[0].value)) if column[0].value else 0
                max_length = max(max_length, title_length)
                
                adjusted_width = (max_length + 2) * 1.2
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 기술평가 전략 시트 추가
            strategy_sheet = workbook.create_sheet("기술평가 전략")
            
            # 스타일 정의
            title_font = Font(name='맑은 고딕', size=14, bold=True)
            header_font = Font(name='맑은 고딕', size=11, bold=True)
            normal_font = Font(name='맑은 고딕', size=10)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 타이틀 추가
            strategy_sheet.merge_cells('A1:F1')
            title_cell = strategy_sheet['A1']
            title_cell.value = "기술적 지표 및 종합평가 방법론"
            title_cell.font = title_font
            title_cell.alignment = center_align
            
            # 1. 종합평가 기준
            strategy_sheet.merge_cells('A3:F3')
            strategy_sheet['A3'].value = "1. 종합평가 기준"
            strategy_sheet['A3'].font = header_font
            
            evaluation_criteria = [
                ['평가', '조건', '설명'],
                ['강매수', '매수 점수가 매도 점수의 2배 이상 & 매수 점수 65% 이상', '다수의 기술적 지표가 강한 매수 신호를 보내는 상태'],
                ['매수', '매수 점수 > 매도 점수 & 매수 점수 45% 이상', '기술적 지표가 매수 우위를 보이는 상태'],
                ['보유', '매수/매도 점수가 균형을 이루는 상태', '뚜렷한 방향성이 없어 현 포지션 유지 권장'],
                ['매도', '매도 점수 > 매수 점수 & 매도 점수 45% 이상', '기술적 지표가 매도 우위를 보이는 상태'],
                ['강매도', '매도 점수가 매수 점수의 2배 이상 & 매도 점수 65% 이상', '다수의 기술적 지표가 강한 매도 신호를 보내는 상태']
            ]
            
            for row_idx, row_data in enumerate(evaluation_criteria, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = strategy_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = header_font if row_idx == 4 else normal_font
                    cell.alignment = center_align if col_idx < 3 else left_align
                    cell.border = thin_border
                    
                    if row_idx > 4:
                        if value == '강매수' or value == '매수':
                            cell.fill = green_fill
                        elif value == '매도' or value == '강매도':
                            cell.fill = red_fill
            
            # 2. 기술적 지표별 해석 및 배점
            strategy_sheet.merge_cells('A10:F10')
            strategy_sheet['A10'].value = "2. 기술적 지표별 해석 및 배점"
            strategy_sheet['A10'].font = header_font
            
            indicator_explanations = [
                ['지표', '매수 신호', '중립 신호', '매도 신호', '가중치', '특징'],
                ['볼린저 밴드', '위치 < 20%', '20% < 위치 < 80%', '위치 > 80%', '2', '변동성과 추세의 강도를 측정'],
                ['MACD', 'MACD > Signal & 상승', 'MACD ≈ Signal', 'MACD < Signal & 하락', '2', '중장기 추세 방향을 파악'],
                ['RSI', 'RSI < 30', '30 < RSI < 70', 'RSI > 70', '2', '과매수/과매도 상태와 모멘텀 측정'],
                ['CCI', 'CCI < -100', '-100 < CCI < 100', 'CCI > 100', '1', '평균 가격과의 괴리 측정'],
                ['ADX', 'ADX > 25 & +DI > -DI', 'ADX < 20', 'ADX > 25 & +DI < -DI', '2', '추세의 강도를 측정'],
                ['Williams %R', 'Williams < -80', '-80 < Williams < -20', 'Williams > -20', '1', '과매수/과매도 상태 측정'],
                ['Stochastic', '%K < 20 & %K > %D', '%K와 %D 교차', '%K > 80 & %K < %D', '2', '모멘텀과 추세 반전 신호 포착']
            ]
            
            for row_idx, row_data in enumerate(indicator_explanations, 11):
                for col_idx, value in enumerate(row_data, 1):
                    cell = strategy_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = header_font if row_idx == 11 else normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # 신호에 따른 색상 지정
                    if row_idx > 11:
                        if col_idx == 2:  # 매수 신호
                            cell.fill = green_fill
                        elif col_idx == 4:  # 매도 신호
                            cell.fill = red_fill
            
            # 3. 지표 간 상관관계 활용 전략
            strategy_sheet.merge_cells('A20:F20')
            strategy_sheet['A20'].value = "3. 지표 간 상관관계 활용 전략"
            strategy_sheet['A20'].font = header_font
            
            synergy_explanations = [
                ['조합', '시너지 효과', '해석'],
                ['MACD + ADX', '추세 방향 + 강도 결합', 'MACD 상승 + ADX > 25 + +DI > -DI = 강한 상승 추세 확인'],
                ['RSI + Williams %R', '모멘텀 지표 교차 검증', '두 지표가 동시에 과매수/과매도 신호시 신뢰도 상승'],
                ['볼린저 밴드 + Stochastic', '변동성 + 모멘텀 결합', '밴드 하단 접근 + Stochastic 상향돌파 = 강한 반등 신호'],
                ['CCI + Stochastic', '사이클 + 반전 신호 결합', '두 지표 모두 과매도에서 반등 신호시 매수 신뢰도 상승']
            ]
            
            for row_idx, row_data in enumerate(synergy_explanations, 21):
                for col_idx, value in enumerate(row_data, 1):
                    cell = strategy_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = header_font if row_idx == 21 else normal_font
                    cell.alignment = center_align if col_idx < 3 else left_align
                    cell.border = thin_border
            
            # 4. 실전 투자 전략
            strategy_sheet.merge_cells('A27:F27')
            strategy_sheet['A27'].value = "4. 실전 투자 전략"
            strategy_sheet['A27'].font = header_font
            
            practical_strategies = [
                ['시장 상황', '활용 지표', '대응 전략'],
                ['상승 추세', 'MACD, ADX, Stochastic', '매수 포지션 강화, 추세 추종 전략 구사'],
                ['하락 추세', 'MACD, ADX, Bollinger', '매도 포지션 강화, 추세 추종 또는 반등 대비'],
                ['횡보장', 'RSI, CCI, Williams %R', '박스권 돌파/이탈 확인 후 대응, 단기 스윙 전략'],
                ['변동성 높음', 'Bollinger 밴드폭, ADX', '포지션 축소, 리스크 관리 강화'],
                ['변동성 낮음', 'Bollinger 밴드폭, MACD', '변동성 확대 대비, 브레이크아웃 전략']
            ]
            
            for row_idx, row_data in enumerate(practical_strategies, 28):
                for col_idx, value in enumerate(row_data, 1):
                    cell = strategy_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = header_font if row_idx == 28 else normal_font
                    cell.alignment = center_align
                    cell.border = thin_border
            
            # 5. 종합평가 해석 가이드
            strategy_sheet.merge_cells('A35:F35')
            strategy_sheet['A35'].value = "5. 종합평가 해석 가이드"
            strategy_sheet['A35'].font = header_font
            
            guide_content = [
                "* 종합평가는 7개 주요 기술적 지표의 신호를 종합적으로 분석한 결과입니다.",
                "* 평가 결과 옆의 괄호 안 숫자 (예: 5:2:3)는 각각 매수:중립:매도 점수를 의미합니다.",
                "* 평가 근거가 있는 경우 '매수 (5:1:2) - 모멘텀 지표 매수신호' 형태로 표시됩니다.",
                "* 모든 기술적 지표는 서로 다른 측면을 측정하므로, 종합적인 판단이 중요합니다.",
                "* 시장 상황과 개별 종목의 펀더멘털을 함께 고려하여 최종 투자 판단을 내려야 합니다.",
                "* 기술적 분석은 과거 데이터 기반의 확률적 접근법으로, 미래 가격을 보장하지 않습니다."
            ]
            
            for row_idx, content in enumerate(guide_content, 36):
                cell = strategy_sheet.cell(row=row_idx, column=1)
                cell.value = content
                cell.font = normal_font
                cell.alignment = left_align
                strategy_sheet.merge_cells(f'A{row_idx}:F{row_idx}')
            
            # 열 너비 자동 조정
            for idx, column in enumerate(strategy_sheet.columns, 1):
                max_length = 0
                column_letter = None
                
                # 각 셀의 내용 길이 확인하여 최대 길이 찾기
                for cell in column:
                    try:
                        # merged cell 건너뛰기
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                                
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # 열 너비 설정 (약간의 여유 공간 추가)
                if column_letter:
                    adjusted_width = (max_length + 2) * 1.0
                    strategy_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 특정 열 너비 수동 조정
            strategy_sheet.column_dimensions['A'].width = 15
            strategy_sheet.column_dimensions['C'].width = 20
            strategy_sheet.column_dimensions['D'].width = 20
            strategy_sheet.column_dimensions['F'].width = 35
        
        # CSV 파일로도 저장
        df.to_csv(csv_filename, index=False, encoding='utf-8-sig', float_format='%.2f')
        
        print(f"\n분석 결과가 저장되었습니다:")
        print(f"- Excel: {excel_filename}")
        print(f"- CSV: {csv_filename}")
        
    except Exception as e:
        print(f"\n엑셀 파일 저장 중 오류 발생: {e}")


if __name__ == "__main__":
    # 종목 정보 읽기
    stocks = read_stocks_file()
    
    if not stocks:
        print("오류: 종목 정보를 읽을 수 없거나 파일이 비어 있습니다.")
    else:
        # 포트폴리오 분석
        results, summary = analyze_portfolio(stocks)
        
        # 표 형식으로 결과 출력 및 엑셀 파일 저장
        display_results_table(results, summary) 